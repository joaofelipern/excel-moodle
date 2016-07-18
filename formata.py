import openpyxl
import csv

wb = openpyxl.load_workbook('/home/nacer/Dropbox/scripts/participantes_curso_ead.xlsx')


sheet = wb.get_sheet_by_name(wb.get_sheet_names().pop())
print sheet

def username(cpf):
    #print type(cpf) 
    cpf = cpf.replace(".","")
    cpf = cpf.replace("-","")
    cpf = cpf.replace(" ","")
    if len(cpf) < 11:         
        cpf = cpf.rjust(11, '0')        
    return cpf

#print sheet.cell(row=1,column=2).value

for coluna in range(1, sheet.get_highest_column()):           
    nomecoluna = raw_input('Renomeie a coluna <' + sheet.cell(row=1,column=coluna).value + '> ou ENTER para continuar: ')
    if nomecoluna != '':    
        sheet.cell(row=1,column=coluna).value = nomecoluna
    else:
        nomecoluna = sheet.cell(row=1,column=coluna).value

    for rowNum in range(2, sheet.max_row+1):  # skip the first row
        valor = sheet.cell(row=rowNum, column=coluna).value
        if nomecoluna == u'username':
            valor = sheet.cell(row=rowNum, column=coluna).value = username(sheet.cell(row=rowNum, column=coluna).value)            
        
        print valor        
    
    print '\n'
        
        
            
    
#    if produceName in PRICE_UPDATES:
#        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('/home/nacer/Dropbox/scripts/participantes_curso_ead.xlsx')
print 'Tabela salva com sucesso!'


sh = wb.get_active_sheet()
with open('test.csv', 'wb') as f:
    c = csv.writer(f)
    for r in sh.rows:
        c.writerow([cell.value.encode('utf-8') for cell in r])


print 'Arquivo test.csv salvo com sucesso!'